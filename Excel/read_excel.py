import openpyxl as xl

wb = xl.Workbook()
ws = wb.active
print(ws)
print(ws['A1'])
ws['A1'] = 520

ws['A3'] = 'hello'

import datetime
ws['A5'] = datetime.datetime.now()
wb.save('demo.xlsx')